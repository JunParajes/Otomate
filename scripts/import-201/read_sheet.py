"""
Read the 201 spreadsheet into records, applying the rules in mapping.py.

Reading is separated from writing on purpose: this module never touches the API
or the database, so it can be run against the real file as often as you like
with no possibility of changing anything.

WHAT THIS FILE ASSUMES ABOUT THE SHEET, all of it checked before it was written:

  * Row 3 is the header. Rows below it are people, except rows where only column
    A or B is filled — those are section labels (a company, "REGULAR EMPLOYEES",
    "SEPARATED EMLOYEE'S") whose meaning applies to every row beneath them.
  * Text dates are MM/DD/YYYY. Verified across all three date columns: 261 cells
    have a second component above 12 and NOT ONE has a first component above 12.
    A dd/mm reading is therefore impossible, not merely less likely.
  * Phone numbers stored as numbers have lost a leading zero — Excel does that to
    anything that looks numeric. A 10-digit value is a mobile number missing its
    0, which is why prepending one is a repair and not a guess.
"""
import datetime
import re
import unicodedata

import openpyxl

import mapping

HEADER_ROW = 3
COL = {
    'no': 1, 'surname': 2, 'first': 3, 'middle': 4,
    'birth': 5, 'birth_yr': 6, 'age': 7,
    'hired': 8, 'ended': 9, 'end_yr': 10, 'service': 11,
    'gender': 12, 'birthplace': 13, 'address': 14, 'phone': 15,
    'position': 16, 'civil': 17, 'status': 18,
    'probationary': 19, 'extend': 20, 'regular': 21,
    'branch': 22, 'status2': 23,
    'sss': 24, 'hdmf': 25, 'philhealth': 26, 'tin': 27,
    'confidentiality': 28, 'authority': 29, 'birth_cert': 30, 'marriage': 31,
    'height': 32, 'weight': 33, 'religion': 34, 'education': 35,
    'email': 36, 'emergency_name': 37, 'emergency_phone': 38, 'remarks': 39,
}

SECTIONS = {'REGULAR EMPLOYEES', 'PROBATIONARY EMPLOYEES', 'EXTRA'}


def _text(v):
    """A trimmed string, or None. Collapses the whitespace Excel leaves behind."""
    if v is None:
        return None
    s = unicodedata.normalize('NFKC', str(v)).replace(' ', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s or None


def _is_nothing(s):
    """True when a cell says 'we do not have this' rather than holding a value."""
    return s is None or s.strip().lower().rstrip('.') in mapping.NOT_A_VALUE


# Generational suffixes, and ONLY these.
#
# The spreadsheet has one Surname column, so "Cruz Jr." was landing whole in the
# surname and the suffix field stayed empty — the name then reads wrong on a COE
# and sorts under the wrong person.
#
# Deliberately a closed list rather than "the last word". Plenty of real
# surnames here are two words — Dela Pena, Dela Torre, San Juan — and plenty of
# first names are two given names, Mae and Ann and Joy being the common ones.
# Splitting on the last token would mangle sixty-odd names to fix nine.
SUFFIXES = {
    'jr': 'Jr.', 'jr.': 'Jr.', 'jr,': 'Jr.', 'junior': 'Jr.',
    'sr': 'Sr.', 'sr.': 'Sr.', 'senior': 'Sr.',
    'ii': 'II', 'iii': 'III', 'iv': 'IV',
}


def split_suffix(name):
    """
    ("Cruz Jr.") -> ("Cruz", "Jr.").  ("Dela Pena") -> ("Dela Pena", None).

    Punctuation is normalised on the way out, so a stray "Jr," becomes "Jr."
    rather than a second spelling of the same thing.
    """
    if not name:
        return name, None
    parts = name.split()
    if len(parts) < 2:
        return name, None
    suffix = SUFFIXES.get(parts[-1].lower())
    if not suffix:
        return name, None
    return ' '.join(parts[:-1]), suffix


class Issue:
    """Something a person has to look at, tied to the cell it came from."""

    def __init__(self, row, column, kind, detail):
        self.row, self.column, self.kind, self.detail = row, column, kind, detail


# ─── Value parsers ───────────────────────────────────────────────────────────

def parse_date(v, row, column, issues, notes=None):
    """
    A date, or None with an issue recorded.

    Real Excel dates are taken as they are. Text is read as MM/DD/YYYY — see the
    module docstring for why that is a fact about this file, not a preference.

    A cell may carry a date AND a word: "Rehired 07/24/2025", "Offially in LDB
    10/01/2025". The date in those is perfectly good and there are 14 of them, so
    it is taken and the word is kept in Remarks. Throwing away a date because
    somebody annotated it would lose real information to tidiness.
    """
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = _text(v)
    if _is_nothing(s):
        return None
    m = re.fullmatch(r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})', s)
    if not m:
        # Exactly one date embedded in other words: take it, keep the words.
        found = re.findall(r'(\d{1,2})\s*[/\-.]\s*(\d{1,2})\s*[/\-.]\s*(\d{2,4})', s)
        if len(found) == 1:
            m = re.match(r'(\d+)\D+(\d+)\D+(\d+)', ''.join(f'{g} ' for g in found[0]))
            if notes is not None:
                notes.append(f'{column} in the sheet: {s}')
        else:
            issues.append(Issue(row, column, 'date-unreadable', s))
            return None
    month, day, year = (int(g) for g in m.groups())
    if year < 100:
        year += 2000 if year < 50 else 1900
    try:
        return datetime.date(year, month, day)
    except ValueError:
        issues.append(Issue(row, column, 'date-impossible', s))
        return None


def parse_phone(v, row, column, issues):
    """
    A Philippine mobile as 11 digits.

    Excel stripped the leading zero from anything it read as a number, so 10
    digits starting 9 is the common case and gets its 0 back. Anything else is
    reported rather than padded into looking valid.
    """
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = _text(v)
    if _is_nothing(s):
        return None
    digits = re.sub(r'\D', '', s)
    if digits.startswith('63') and len(digits) == 12:
        digits = '0' + digits[2:]
    if len(digits) == 10 and digits.startswith('9'):
        digits = '0' + digits
    if len(digits) == 11 and digits.startswith('09'):
        return digits
    issues.append(Issue(row, column, 'phone-unreadable', s))
    return None


def parse_height_cm(v, row, issues):
    """Feet and inches ("5'7") to whole centimetres, which is what is stored."""
    s = _text(v)
    if _is_nothing(s):
        return None
    m = re.fullmatch(r"(\d)\s*['’;:/]\s*(\d{1,2})\"?", s)
    if not m:
        issues.append(Issue(row, 'height', 'height-unreadable', s))
        return None
    feet, inches = int(m.group(1)), int(m.group(2))
    if inches > 11:
        issues.append(Issue(row, 'height', 'height-impossible', s))
        return None
    return round((feet * 12 + inches) * 2.54)


def parse_weight_grams(v, row, issues):
    """Kilograms to grams — no float ever reaches the record."""
    s = _text(v)
    if _is_nothing(s):
        return None
    try:
        kg = float(re.sub(r'[^\d.]', '', s))
    except ValueError:
        issues.append(Issue(row, 'weight', 'weight-unreadable', s))
        return None
    if not 25 <= kg <= 250:
        issues.append(Issue(row, 'weight', 'weight-implausible', s))
        return None
    return round(kg * 1000)


def parse_education(v, row, issues):
    """(level, detail). The level is matched on the longest key that appears."""
    s = _text(v)
    if _is_nothing(s):
        return None, None
    low = s.lower()
    for key, level in sorted(mapping.EDUCATION_LEVELS, key=lambda p: -len(p[0])):
        if key in low:
            # Whatever the cell said beyond the level itself is worth keeping.
            detail = re.sub(re.escape(key), '', low, flags=re.I)
            # Separators first: an underscore is a word character, so
            # "Secondary_High" hides "secondary" from a \b-anchored match.
            detail = re.sub(r'[_()\[\],]+', ' ', detail)
            # Words that only restate the level. "Senior High School (TVL)"
            # should keep TVL and nothing else — the level already says the rest,
            # and "Secondary" is just another word for high school.
            detail = re.sub(r'\b(graduate|grad|level|undergrad(uate)?|school|with|secondary)\b',
                            '', detail, flags=re.I)
            detail = re.sub(r'[^\w\s&/.-]', ' ', detail)
            detail = re.sub(r'\s+', ' ', detail).strip(' /-._')
            return level, (detail.upper() if detail and len(detail) <= 60 else None)
    issues.append(Issue(row, 'education', 'education-unmapped', s))
    return None, None


def parse_id_number(v, row, column, issues):
    """
    A government ID, kept EXACTLY as typed.

    Otomate stores these unnormalised on purpose — the punctuation is how the
    number appears on the filings. PhilHealth turns up grouped both 2-9-1 and
    4-4-4; both are the same twelve digits and both are correct.
    """
    s = _text(v)
    if _is_nothing(s):
        return None, s if s else None  # value, word-worth-keeping
    if not re.search(r'\d', s):
        issues.append(Issue(row, column, 'id-has-no-digits', s))
        return None, s
    return s, None


def parse_document(v, row, column, issues):
    """MISSING / ON_FILE / NOT_APPLICABLE, or None when the cell is a note."""
    s = _text(v)
    if s is None:
        return None, None
    status = mapping.DOCUMENT_STATUS.get(s.lower())
    if status is None:
        issues.append(Issue(row, column, 'document-unmapped', s))
        return None, s
    return status, None


# ─── The sheet ───────────────────────────────────────────────────────────────

def read(path):
    """
    Returns (people, issues, skipped).

    `people` are dicts of parsed values plus the bookkeeping needed to import
    them; `issues` is everything a person should look at; `skipped` are rows
    deliberately not imported, with the reason.
    """
    ws = openpyxl.load_workbook(path, data_only=True)['Details']
    issues, people, skipped = [], [], []
    company = section = None
    separated_block = False

    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        vals = {name: ws.cell(r, c).value for name, c in COL.items()}
        filled = [k for k, v in vals.items() if v not in (None, '')]
        if not filled:
            continue

        # A label row: only column A or B, and it names a company or a section.
        lone = [k for k in filled if COL[k] <= 2]
        if len(filled) <= 2 and lone:
            label = (_text(vals[lone[0]]) or '').strip()
            up = label.upper()
            if 'SEPARATED' in up:
                separated_block = True
                continue
            if label.lower() in mapping.COMPANIES or label.lower() in mapping.COMPANY_CLOSED:
                company = label.lower()
                continue
            if up in SECTIONS or up.startswith('EXTRA'):
                section = up
                continue
            skipped.append((r, f'stray label {label!r} — not a person, not a known section'))
            continue

        surname, first = _text(vals['surname']), _text(vals['first'])
        if not surname and not first:
            skipped.append((r, 'no name in the row'))
            continue

        if company in mapping.COMPANY_CLOSED:
            skipped.append((r, 'Chicken and Freshness — closed, not imported'))
            continue

        people.append(_person(r, vals, company, section, separated_block, issues))

    return people, issues, skipped


def _person(r, vals, company, section, separated_block, issues):
    notes = []  # words rescued from cells that could not become values

    # ─── Status, which decides almost everything else ───────────────────────
    status_word = (_text(vals['status']) or '').lower()
    status2_word = (_text(vals['status2']) or '').lower()

    employment_type = mapping.EMPLOYMENT_TYPE.get(status_word)
    if employment_type is None and status_word and status_word not in mapping.EXIT_WORDS:
        issues.append(Issue(r, 'status', 'status-unmapped', status_word))

    """
    Which section the row sits in decides, and column W only decides within it.

    Column W was the plan, but the file disagrees. Sixteen rows below the
    SEPARATED header still say "Active" — and every one of them has a real end
    date AFTER its hire date (hired 03/01, ended 04/28), two of them having been
    rehired and left a second time. They have left; the word is simply stale.

    Moving a row into the archive is a deliberate act. Leaving a cell reading
    "Active" is what happens when nobody edits it. So the section wins.

    Within the active block, W still decides — it correctly catches the two
    people marked AWOL and END who were never moved down.
    """
    is_active = mapping.ACTIVE.get(status2_word)
    if is_active is None and status2_word:
        issues.append(Issue(r, 'status2', 'status2-unmapped', status2_word))
    if separated_block:
        if is_active:
            issues.append(Issue(r, 'status2', 'stale-active-in-the-archive', status2_word))
        is_active = False
    elif is_active is None:
        is_active = True

    ended = parse_date(vals['ended'], r, 'ended', issues, notes)
    reason = mapping.SEPARATION_REASON.get(status2_word) or mapping.SEPARATION_REASON.get(status_word)
    # Somebody who went AWOL has no last day by definition — they stopped coming
    # and nobody wrote a date. Only a resignation or an ended contract should
    # have one, so only those are worth reporting as missing.
    if not is_active and not ended and reason != 'AWOL':
        issues.append(Issue(r, 'ended', 'separated-without-a-last-day', status2_word or '(blank)'))

    # Somebody marked as having left while still filed under an active section.
    # Rarer and the opposite problem, so it keeps its own name.
    if not is_active and not separated_block:
        issues.append(Issue(r, 'status2', 'separated-but-filed-under-active', status2_word))

    # ─── Government IDs — the word is kept when the number is not ───────────
    ids = {}
    for key, label in (('sss', 'SSS'), ('hdmf', 'Pag-IBIG'),
                       ('philhealth', 'PhilHealth'), ('tin', 'TIN')):
        value, word = parse_id_number(vals[key], r, key, issues)
        ids[key] = value
        if word:
            notes.append(f'{label}: {word}')

    # ─── Paperwork ─────────────────────────────────────────────────────────
    docs = {}
    for key, label in (('confidentiality', 'Confidentiality agreement'),
                       ('authority', 'Authority to deduct'),
                       ('birth_cert', 'Birth certificate'),
                       ('marriage', 'Marriage contract')):
        status, word = parse_document(vals[key], r, key, issues)
        docs[key] = status
        if word:
            notes.append(f'{label}: {word}')

    branch_word = (_text(vals['branch']) or '').lower()
    branch = mapping.BRANCHES.get(branch_word)
    if branch is None and branch_word:
        issues.append(Issue(r, 'branch', 'branch-unmapped', branch_word))
        if branch_word in mapping.BRANCH_UNRESOLVED:
            notes.append(f'Branch in the sheet: {_text(vals["branch"])}')

    level, detail = parse_education(vals['education'], r, issues)

    # The position column is NOT mapped — the naming is still being settled, so
    # everyone imports as Unassigned and the sheet's word is kept to sort by.
    position_word = _text(vals['position'])
    if position_word:
        notes.append(f'Position in the sheet: {position_word}')

    # The sheet has no suffix column, so "Jr." rides along on whichever name it
    # was typed into — usually the surname, twice the first name.
    last_name, suffix = split_suffix(_text(vals['surname']))
    first_name, first_suffix = split_suffix(_text(vals['first']))
    suffix = suffix or first_suffix

    existing_remarks = _text(vals['remarks'])
    if existing_remarks:
        notes.insert(0, existing_remarks)

    person = {
        'row': r,
        'company': mapping.COMPANIES.get(company or ''),
        'section': section,
        'lastName': last_name,
        'firstName': first_name,
        'middleName': _text(vals['middle']),
        'suffix': suffix,
        'birthDate': parse_date(vals['birth'], r, 'birth', issues, notes),
        'birthPlace': _text(vals['birthplace']),
        'gender': mapping.GENDER.get((_text(vals['gender']) or '').lower()),
        'civilStatus': mapping.CIVIL_STATUS.get((_text(vals['civil']) or '').lower()),
        'religion': None if _is_nothing(_text(vals['religion'])) else _text(vals['religion']),
        'address': None if _is_nothing(_text(vals['address'])) else _text(vals['address']),
        'email': _email(vals['email'], r, issues),
        'heightCm': parse_height_cm(vals['height'], r, issues),
        'weightGrams': parse_weight_grams(vals['weight'], r, issues),
        'educationLevel': level,
        'educationDetail': detail,
        'phone': parse_phone(vals['phone'], r, 'phone', issues),
        'emergencyName': None if _is_nothing(_text(vals['emergency_name'])) else _text(vals['emergency_name']),
        'emergencyContact': parse_phone(vals['emergency_phone'], r, 'emergency_phone', issues),
        'dateHired': parse_date(vals['hired'], r, 'hired', issues, notes),
        'employmentType': employment_type,
        'isActive': is_active,
        'separatedAt': ended,
        'separationReason': reason,
        'branch': branch,
        'remarks': ' · '.join(notes) or None,
        **ids,
        'documents': docs,
    }
    _checksum(person, vals, r, issues)
    return person


def _email(v, r, issues):
    s = _text(v)
    if _is_nothing(s):
        return None
    if not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', s):
        issues.append(Issue(r, 'email', 'email-unreadable', s))
        return None
    return s.lower()


def _checksum(person, vals, r, issues):
    """
    The sheet's own derived columns, used as a check on the dates.

    This is the point of the redundancy: if the typed birth year disagrees with
    the year in the birth date, one of the two is a typo — and a wrong birth date
    is invisible once it is in the system. Same for the age column.
    """
    birth = person['birthDate']
    yr = vals['birth_yr']
    if birth and isinstance(yr, (int, float)) and int(yr) != birth.year:
        issues.append(Issue(r, 'birth', 'birth-year-disagrees',
                            f'date says {birth.year}, Yr column says {int(yr)}'))
    age = vals['age']
    if birth and isinstance(age, (int, float)):
        """
        The Age column is STALE, not wrong, and that is not an error.

        It was typed when each person was added and never recomputed, so across a
        sheet covering 2024-2026 the offsets smear: 161 rows agree with 2026, 71
        are a year behind, 40 two, 18 three, 5 four. Flagging those as
        disagreements produced 68 false alarms and buried the five that matter.

        So only an offset outside that drift is reported. The real ones are
        wild — 10, 24, 25 and 35 years out — and those are typos.
        """
        drift = int(age) - (2026 - birth.year)
        if not -5 <= drift <= 1:
            issues.append(Issue(r, 'age', 'age-disagrees',
                                f'birth year {birth.year} implies about {2026 - birth.year}, '
                                f'Age column says {int(age)}'))
    if birth and birth.year >= 2009:
        issues.append(Issue(r, 'birth', 'under-18', f'born {birth.year}'))
    hired, ended = person['dateHired'], person['separatedAt']
    if hired and ended and ended < hired:
        issues.append(Issue(r, 'ended', 'left-before-hired', f'{hired} → {ended}'))
